from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from urllib.parse import urlparse
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

wait_duration = 10
chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)
driver.maximize_window()

username = "atharvachoudhari.truefunds@gmail.com"
password = "atharva@123"
login_url = "https://www.valueresearchonline.com/funds/17366/quant-small-cap-fund-direct-plan/#other"

driver.get(login_url)

email_input = WebDriverWait(driver, wait_duration).until(
    EC.visibility_of_element_located((By.ID, "user_email"))
)
email_input.send_keys(username)

password_input = WebDriverWait(driver, wait_duration).until(
    EC.visibility_of_element_located((By.ID, "user_password"))
)
password_input.send_keys(password)

login_button = WebDriverWait(driver, wait_duration).until(
    EC.element_to_be_clickable((By.NAME, "commit"))
)
login_button.click()

urls = [
    "https://www.valueresearchonline.com/funds/2310/icici-prudential-value-discovery-fund/",
]

excel_filename = "combined_data.xlsx"
wb = load_workbook(excel_filename)

for url in urls:
    try:
        driver.get(url + "#other")

        parent_div_xpath_h3 = (
            "/html/body/section[2]/div[2]/div/div[7]/section/div/div[2]/div/div"
        )

        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_h3))
        )

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        if fund_name in wb.sheetnames:
            ws = wb[fund_name]
            start_column = ws.max_column + 1 if ws.max_column else 1
        else:
            ws = wb.create_sheet(fund_name)
            start_column = 1

        for i, data in enumerate(combined_data_li, start=1):
            ws.cell(row=i, column=start_column, value=data)

        print(f"Data added to '{fund_name}' sheet in {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

wb.save(excel_filename)

driver.quit()
