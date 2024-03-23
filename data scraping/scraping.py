from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from urllib.parse import urlparse
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

######################################################################################################
# Define wait_duration
wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)

# portfolio


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/2310/icici-prudential-value-discovery-fund/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for index, url in enumerate(urls):
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        combined_df_temp["Fund Name"] = fund_name

        # Saving data to Excel
        excel_filename = "combined_data.xlsx"
        if index == 0:
            combined_df_temp.to_excel(excel_filename, sheet_name=fund_name, index=False)
        else:
            with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)

        print(f"All Data saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

######################################################################################################


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from urllib.parse import urlparse
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Define wait_duration
wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/633/sbi-contra-fund/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

######################################################################################################

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from urllib.parse import urlparse
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Define wait_duration
wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/66/quant-small-cap-fund",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

######################################################################################################

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from urllib.parse import urlparse
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Define wait_duration
wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/22334/axis-small-cap-fund-regular-plan/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

######################################################################################################

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from urllib.parse import urlparse
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Define wait_duration
wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/2597/kotak-small-cap-fund-regular-plan/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

######################################################################################################


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from urllib.parse import urlparse
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Define wait_duration
wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/11463/nippon-india-small-cap-fund/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()
######################################################################################################


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from urllib.parse import urlparse
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Define wait_duration
wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/952/quant-active-fund/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()
######################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/958/quant-mid-cap-fund/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()
######################################################################################################


# Define wait_duration
wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/12052/axis-midcap-fund/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################

# Define wait_duration
wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/4270/kotak-emerging-equity-fund-regular-plan/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


# Define wait_duration
wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/183/nippon-india-growth-fund/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


# Define wait_duration
wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/2662/sbi-magnum-midcap-fund/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from urllib.parse import urlparse
from openpyxl import Workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Define wait_duration
wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


url = "https://www.valueresearchonline.com/funds/103/tata-midcap-growth-fund-regular-plan/"

try:
    driver.get(url)

    # Define XPaths
    parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
    parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
    parent_div_xpath_li = (
        "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
    )

    # Wait for visibility of elements
    WebDriverWait(driver, wait_duration).until(
        EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
    )

    parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
    list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
    combined_data_p = [item.text for item in list_items_p]

    parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
    list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
    combined_data_h3 = [h3.text for h3 in list_items_h3]

    parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
    list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
    combined_data_li = [item.text for item in list_items_li]

    all_table_dfs = process_tables(url)

    max_len = max(
        len(combined_data_h3),
        len(combined_data_p),
        len(combined_data_li),
        *[len(df) for df in all_table_dfs],
    )

    combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
    combined_data_p.extend([""] * (max_len - len(combined_data_p)))
    combined_data_li.extend([""] * (max_len - len(combined_data_li)))

    combined_df_temp = pd.DataFrame(
        {
            "List Items (h3)": combined_data_h3,
            "List Items (p)": combined_data_p,
            "List Items (li)": combined_data_li,
        }
    )

    for df in all_table_dfs:
        for col in df.columns:
            if col in combined_df_temp.columns:
                combined_df_temp[col] = (
                    combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                )
                combined_df_temp[col] = combined_df_temp[col].str.replace(
                    r"(?:^|\n)nan(?:$|\n)", "\n", regex=True
                )
            else:
                combined_df_temp[col] = df[col]

    parsed_url = urlparse(url)
    fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

    # Saving data to Excel with fund name as sheet name
    excel_filename = "combined_data.xlsx"
    with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
        workbook = writer.book
        try:
            # Try to get the existing workbook
            workbook = load_workbook(excel_filename)
        except FileNotFoundError:
            # If the file does not exist, create a new workbook
            workbook = Workbook()
        finally:
            # Add the DataFrame to a new sheet
            combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
            print(f"All Data for {fund_name} saved to {excel_filename}")

except Exception as e:
    print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()


############################################################################################################################

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from urllib.parse import urlparse
from openpyxl import Workbook, load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Define wait_duration
wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


url = "https://www.valueresearchonline.com/funds/509/tata-large-cap-fund-regular-plan/#fund-portfolio"

try:
    driver.get(url)

    # Define XPaths
    parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
    parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
    parent_div_xpath_li = (
        "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
    )

    # Wait for visibility of elements
    WebDriverWait(driver, wait_duration).until(
        EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
    )

    parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
    list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
    combined_data_p = [item.text for item in list_items_p]

    parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
    list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
    combined_data_h3 = [h3.text for h3 in list_items_h3]

    parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
    list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
    combined_data_li = [item.text for item in list_items_li]

    all_table_dfs = process_tables(url)

    max_len = max(
        len(combined_data_h3),
        len(combined_data_p),
        len(combined_data_li),
        *[len(df) for df in all_table_dfs],
    )

    combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
    combined_data_p.extend([""] * (max_len - len(combined_data_p)))
    combined_data_li.extend([""] * (max_len - len(combined_data_li)))

    combined_df_temp = pd.DataFrame(
        {
            "List Items (h3)": combined_data_h3,
            "List Items (p)": combined_data_p,
            "List Items (li)": combined_data_li,
        }
    )

    for df in all_table_dfs:
        for col in df.columns:
            if col in combined_df_temp.columns:
                combined_df_temp[col] = (
                    combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                )
                combined_df_temp[col] = combined_df_temp[col].str.replace(
                    r"(?:^|\n)nan(?:$|\n)", "\n", regex=True
                )
            else:
                combined_df_temp[col] = df[col]

    parsed_url = urlparse(url)
    fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

    # Saving data to Excel with fund name as sheet name
    excel_filename = "combined_data.xlsx"
    with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
        workbook = writer.book
        try:
            # Try to get the existing workbook
            workbook = load_workbook(excel_filename)
        except FileNotFoundError:
            # If the file does not exist, create a new workbook
            workbook = Workbook()
        finally:
            # Add the DataFrame to a new sheet
            combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
            print(f"All Data for {fund_name} saved to {excel_filename}")

except Exception as e:
    print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()


########################################################################################################################

wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/4871/edelweiss-large-mid-cap-fund-regular-plan/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/3083/sbi-bluechip-fund/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/10780/axis-bluechip-fund/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/5270/nippon-india-large-cap-fund/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/577/kotak-bluechip-fund-regular-plan/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/4111/quant-large-and-mid-cap-fund/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/102/tata-large-mid-cap-fund-regular-plan/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/10432/edelweiss-large-cap-fund-regular-plan/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/197/sbi-large-midcap-fund/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/11333/canara-robeco-bluechip-equity-fund-regular-plan/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/11213/mirae-asset-emerging-bluechip-fund-regular-plan/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/6227/edelweiss-mid-cap-fund-regular-plan/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/19699/parag-parikh-flexi-cap-fund-regular-plan/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/38583/icici-prudential-india-opportunities-fund/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/41614/nippon-india-flexi-cap-fund-regular-plan/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/42365/quant-large-cap-fund-regular-plan/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/3001/baroda-bnp-paribas-elss-fund/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/8216/quant-focused-fund/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/40104/parag-parikh-tax-saver-fund-regular-plan/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/10032/bandhan-elss-tax-saver-fund-regular-plan/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/2937/kotak-elss-tax-saver-regular-plan/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/42502/sundaram-flexi-cap-fund-regular-plan/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/1471/icici-prudential-multi-asset-fund/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################


wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/737/icici-prudential-technology-fund/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        all_table_dfs = process_tables(url)

        max_len = max(
            len(combined_data_h3),
            len(combined_data_p),
            len(combined_data_li),
            *[len(df) for df in all_table_dfs],
        )

        combined_data_h3.extend([""] * (max_len - len(combined_data_h3)))
        combined_data_p.extend([""] * (max_len - len(combined_data_p)))
        combined_data_li.extend([""] * (max_len - len(combined_data_li)))

        combined_df_temp = pd.DataFrame(
            {
                "List Items (h3)": combined_data_h3,
                "List Items (p)": combined_data_p,
                "List Items (li)": combined_data_li,
            }
        )

        for df in all_table_dfs:
            for col in df.columns:
                if col in combined_df_temp.columns:
                    new_col_name = f"{col}_{len(combined_df_temp.columns)}"
                    combined_df_temp[new_col_name] = (
                        combined_df_temp[col].astype(str) + "\n" + df[col].astype(str)
                    )
                    combined_df_temp[new_col_name] = combined_df_temp[
                        new_col_name
                    ].str.replace(r"(?:^|\n)nan(?:$|\n)", "\n", regex=True)
                else:
                    combined_df_temp[col] = df[col]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        # Saving data to Excel with fund name as sheet name
        excel_filename = "combined_data.xlsx"
        with pd.ExcelWriter(excel_filename, engine="openpyxl", mode="a") as writer:
            workbook = writer.book
            try:
                # Try to get the existing workbook
                workbook = load_workbook(excel_filename)
            except FileNotFoundError:
                # If the file does not exist, create a new workbook
                workbook = Workbook()
            finally:
                # Add the DataFrame to a new sheet
                combined_df_temp.to_excel(writer, sheet_name=fund_name, index=False)
                print(f"All Data for {fund_name} saved to {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()

#########################################################################################################################
