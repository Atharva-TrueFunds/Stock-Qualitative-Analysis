from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

wait_duration = 10

urls = [
    "https://www.moneycontrol.com/mutual-funds/icici-prudential-value-discovery-fund/investment-info/MPI087",
    "https://www.moneycontrol.com/mutual-funds/sbi-contra-fund/investment-info/MSB010",
    "https://www.moneycontrol.com/mutual-funds/quant-small-cap-fund-direct-plan/investment-info/MES056",
    "https://www.moneycontrol.com/mutual-funds/axis-small-cap-fund-direct-plan/investment-info/MAA286",
    "https://www.moneycontrol.com/mutual-funds/kotak-small-cap-fund-direct-plan/investment-info/MKC307",
    "https://www.moneycontrol.com/mutual-funds/nippon-india-small-cap-fund-direct-plan/investment-info/MNI001",
    "https://www.moneycontrol.com/mutual-funds/quant-active-fund-direct-plan/investment-info/MQU001",
    "https://www.moneycontrol.com/mutual-funds/quant-mid-cap-fund-direct-plan/investment-info/MQT002",
    "https://www.moneycontrol.com/mutual-funds/axis-mid-cap-fund-direct-plan/investment-info/MAA273",
    "https://www.moneycontrol.com/mutual-funds/kotak-emerging-equity-fund-direct-plan/investment-info/MKC216",
    "https://www.moneycontrol.com/mutual-funds/nippon-india-growth-fund-direct-plan/investment-info/MNI003",
    "https://www.moneycontrol.com/mutual-funds/sbi-midcap-fund-direct-plan/investment-info/MSB260",
    "https://www.moneycontrol.com/mutual-funds/tata-mid-cap-growth-fund-direct-plan/investment-info/MTA210",
    "https://www.moneycontrol.com/mutual-funds/tata-large-cap-fund-direct-plan/investment-info/MTA002",
    "https://www.moneycontrol.com/mutual-funds/edelweiss-large-midcap-fund-direct-plan/investment-info/MAB006",
    "https://www.moneycontrol.com/mutual-funds/sbi-blue-chip-fund-direct-plan/investment-info/MSB105",
    "https://www.moneycontrol.com/mutual-funds/axis-blue-chip-fund-direct-plan/investment-info/MAA287",
    "https://www.moneycontrol.com/mutual-funds/nippon-india-large-cap-fund-direct-plan/investment-info/MNI004",
    "https://www.moneycontrol.com/mutual-funds/kotak-bluechip-fund-direct-plan/investment-info/MKC002",
    "https://www.moneycontrol.com/mutual-funds/quant-large-mid-cap-fund-direct-plan/investment-info/MQT001",
    "https://www.moneycontrol.com/mutual-funds/tata-large-midcap-fund-direct-plan/investment-info/MTA209",
    "https://www.moneycontrol.com/mutual-funds/edelweiss-large-cap-fund-direct-plan/investment-info/MAB004",
    "https://www.moneycontrol.com/mutual-funds/sbi-large-mid-cap-fund-direct-plan/investment-info/MSB263",
    "https://www.moneycontrol.com/mutual-funds/canara-robeco-blue-chip-equity-fund-direct-plan/investment-info/MCR028",
    "https://www.moneycontrol.com/mutual-funds/mirae-asset-emerging-bluechip-fund-direct-plan/investment-info/MMI004",
    "https://www.moneycontrol.com/mutual-funds/edelweiss-mid-cap-fund-direct-plan/investment-info/MAB003",
    "https://www.moneycontrol.com/mutual-funds/parag-parikh-flexi-cap-fund-direct-plan/investment-info/MPP004",
    "https://www.moneycontrol.com/mutual-funds/icici-prudential-india-opportunities-fund-direct-plan/investment-info/MPI050",
    "https://www.moneycontrol.com/mutual-funds/nippon-india-flexicap-fund-direct-plan/investment-info/MNI012",
    "https://www.moneycontrol.com/mutual-funds/quant-large-cap-fund-direct-plan/investment-info/MQT004",
    "https://www.moneycontrol.com/mutual-funds/baroda-bnp-paribas-tax-savings-fund-direct-plan/investment-info/MBN004",
    "https://www.moneycontrol.com/mutual-funds/quant-focused-fund-regular-plan/investment-info/MQU002",
    "https://www.moneycontrol.com/mutual-funds/parag-parikh-tax-saver-fund-regular-plan/investment-info/MPP002",
    "https://www.moneycontrol.com/mutual-funds/bandhan-bank-tax-saver-fund-regular-plan/investment-info/MBB001",
    "https://www.moneycontrol.com/mutual-funds/kotak-tax-saver-fund-regular-plan/investment-info/MKM240",
    "https://www.moneycontrol.com/mutual-funds/sundaram-flexi-cap-fund-regular-plan/investment-info/MSU079",
    "https://www.moneycontrol.com/mutual-funds/icici-prudential-multi-asset-fund-regular-plan/investment-info/MPI204",
    "https://www.moneycontrol.com/mutual-funds/icici-prudential-technology-fund-regular-plan/investment-info/MPI048",
]


chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)
for index, url in enumerate(urls):
    try:
        for index, url in enumerate(urls):
            try:
                driver.get(url)

                parent_div_xpath1 = "/html/body/div[15]/section[3]/div/div/div[1]"

                # Find li elements
                li_elements = driver.find_elements(By.XPATH, f"{parent_div_xpath1}//p")
                combined_data_li = [
                    elem.text.encode("utf-8").decode("ascii", "ignore")
                    for elem in li_elements
                ]
                xpath2 = "/html/body/div[15]/section[3]/div/div/div[1]/div[3]"
                span_elements = driver.find_elements(By.XPATH, f"{xpath2}//li")
                combined_data_span = [
                    elem.text.encode("utf-8").decode("ascii", "ignore")
                    for elem in span_elements
                ]

                print("Data extracted successfully (LI):", combined_data_li)
                print("Data extracted successfully (SPAN):", combined_data_span)

                excel_filename = "money_control.xlsx"
                wb = load_workbook(excel_filename)

                sheet_name = "Quant Small Cap Fund"
                ws = wb[sheet_name]

                start_column = ws.max_column + 1 if ws.max_column else 1

                for row, value in enumerate(combined_data_li, start=1):
                    ws.cell(row=row, column=start_column, value=value)
                for row, value in enumerate(combined_data_span, start=1):
                    ws.cell(row=row, column=start_column, value=value)

                wb.save(excel_filename)
                print(f"Data added to '{sheet_name}' sheet in {excel_filename}")

            except Exception as e:
                print(f"Error processing URL: {url}. Error: {str(e)}")

    finally:
        driver.quit()
