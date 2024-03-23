from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import pandas as pd
from openpyxl import load_workbook
from urllib.parse import urlparse
import time

wait_duration = 5

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/2310/icici-prudential-value-discovery-fund/",
    "https://www.valueresearchonline.com/funds/633/sbi-contra-fund/",
    "https://www.valueresearchonline.com/funds/66/quant-small-cap-fund/",
    "https://www.valueresearchonline.com/funds/22334/axis-small-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/2597/kotak-small-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/11463/nippon-india-small-cap-fund/",
    "https://www.valueresearchonline.com/funds/952/quant-active-fund/",
    "https://www.valueresearchonline.com/funds/958/quant-mid-cap-fund/",
    "https://www.valueresearchonline.com/funds/12052/axis-midcap-fund/",
    "https://www.valueresearchonline.com/funds/4270/kotak-emerging-equity-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/183/nippon-india-growth-fund/",
    "https://www.valueresearchonline.com/funds/2662/sbi-magnum-midcap-fund/",
    "https://www.valueresearchonline.com/funds/103/tata-midcap-growth-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/509/tata-large-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/4871/edelweiss-large-mid-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/3083/sbi-bluechip-fund/",
    "https://www.valueresearchonline.com/funds/10780/axis-bluechip-fund/",
    "https://www.valueresearchonline.com/funds/5270/nippon-india-large-cap-fund/",
    "https://www.valueresearchonline.com/funds/577/kotak-bluechip-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/4111/quant-large-and-mid-cap-fund/",
    "https://www.valueresearchonline.com/funds/102/tata-large-mid-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/10432/edelweiss-large-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/197/sbi-large-midcap-fund/",
    "https://www.valueresearchonline.com/funds/11333/canara-robeco-bluechip-equity-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/11213/mirae-asset-emerging-bluechip-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/6227/edelweiss-mid-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/19699/parag-parikh-flexi-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/38583/icici-prudential-india-opportunities-fund/",
    "https://www.valueresearchonline.com/funds/41614/nippon-india-flexi-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/42365/quant-large-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/3001/baroda-bnp-paribas-elss-fund/",
    "https://www.valueresearchonline.com/funds/8216/quant-focused-fund/",
    "https://www.valueresearchonline.com/funds/40104/parag-parikh-tax-saver-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/10032/bandhan-elss-tax-saver-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/2937/kotak-elss-tax-saver-regular-plan/",
    "https://www.valueresearchonline.com/funds/42502/sundaram-flexi-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/1471/icici-prudential-multi-asset-fund/",
    "https://www.valueresearchonline.com/funds/737/icici-prudential-technology-fund/",
]


for i in range(len(urls)):
    urls[i] = urls[i] + "#performance"

for index, url in enumerate(urls):
    chrome_service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=chrome_service)
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    all_dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            all_dfs.append(df)
    if not all_dfs:
        combined_df_temp = pd.DataFrame({"No Data": [float("nan")]})
    else:
        combined_df_temp = pd.DataFrame()  # Initialize DataFrame
        # Concatenate all DataFrames
        for df_index, df in enumerate(all_dfs):
            for col_index, col in enumerate(df.columns):
                new_col_name = f"{col} - {df_index}"  # Create new column name based on DataFrame index
                combined_df_temp[new_col_name] = df[col]

    # Remove empty columns
    combined_df_temp = combined_df_temp.dropna(axis=1, how="all")

    parsed_url = urlparse(url)
    fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]
    excel_filename = "combined_data.xlsx"

    wb = load_workbook(excel_filename)

    if fund_name in wb.sheetnames:
        ws = wb[fund_name]
        start_column = ws.max_column + 1 if ws.max_column else 1
    else:
        ws = wb.create_sheet(fund_name)

    for col_index, col_name in enumerate(combined_df_temp.columns, start=start_column):
        ws.cell(row=1, column=col_index)  # Write column names
        for row_index, data in enumerate(combined_df_temp[col_name], start=2):
            ws.cell(row=row_index, column=col_index, value=data)  # Write data

    wb.save(excel_filename)
    print(f"Data added to '{fund_name}' sheet in {excel_filename}")

driver.quit()
###################################################################################################

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/2310/icici-prudential-value-discovery-fund/",
    "https://www.valueresearchonline.com/funds/633/sbi-contra-fund/",
    "https://www.valueresearchonline.com/funds/66/quant-small-cap-fund/",
    "https://www.valueresearchonline.com/funds/22334/axis-small-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/2597/kotak-small-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/11463/nippon-india-small-cap-fund/",
    "https://www.valueresearchonline.com/funds/952/quant-active-fund/",
    "https://www.valueresearchonline.com/funds/958/quant-mid-cap-fund/",
    "https://www.valueresearchonline.com/funds/12052/axis-midcap-fund/",
    "https://www.valueresearchonline.com/funds/4270/kotak-emerging-equity-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/183/nippon-india-growth-fund/",
    "https://www.valueresearchonline.com/funds/2662/sbi-magnum-midcap-fund/",
    "https://www.valueresearchonline.com/funds/103/tata-midcap-growth-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/509/tata-large-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/4871/edelweiss-large-mid-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/3083/sbi-bluechip-fund/",
    "https://www.valueresearchonline.com/funds/10780/axis-bluechip-fund/",
    "https://www.valueresearchonline.com/funds/5270/nippon-india-large-cap-fund/",
    "https://www.valueresearchonline.com/funds/577/kotak-bluechip-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/4111/quant-large-and-mid-cap-fund/",
    "https://www.valueresearchonline.com/funds/102/tata-large-mid-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/10432/edelweiss-large-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/197/sbi-large-midcap-fund/",
    "https://www.valueresearchonline.com/funds/11333/canara-robeco-bluechip-equity-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/11213/mirae-asset-emerging-bluechip-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/6227/edelweiss-mid-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/19699/parag-parikh-flexi-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/38583/icici-prudential-india-opportunities-fund/",
    "https://www.valueresearchonline.com/funds/41614/nippon-india-flexi-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/42365/quant-large-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/3001/baroda-bnp-paribas-elss-fund/",
    "https://www.valueresearchonline.com/funds/8216/quant-focused-fund/",
    "https://www.valueresearchonline.com/funds/40104/parag-parikh-tax-saver-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/10032/bandhan-elss-tax-saver-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/2937/kotak-elss-tax-saver-regular-plan/",
    "https://www.valueresearchonline.com/funds/42502/sundaram-flexi-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/1471/icici-prudential-multi-asset-fund/",
    "https://www.valueresearchonline.com/funds/737/icici-prudential-technology-fund/",
]


for i in range(len(urls)):
    urls[i] = urls[i] + "#risk"

for index, url in enumerate(urls):
    chrome_service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=chrome_service)
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    all_dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            all_dfs.append(df)

    combined_df_temp = pd.DataFrame()  # Initialize DataFrame
    # Concatenate all DataFrames
    for df_index, df in enumerate(all_dfs):
        for col_index, col in enumerate(df.columns):
            new_col_name = (
                f"{col} - {df_index}"  # Create new column name based on DataFrame index
            )
            combined_df_temp[new_col_name] = df[col]

    # Remove empty columns
    combined_df_temp = combined_df_temp.dropna(axis=1, how="all")

    parsed_url = urlparse(url)
    fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]
    excel_filename = "combined_data.xlsx"

    wb = load_workbook(excel_filename)

    if fund_name in wb.sheetnames:
        ws = wb[fund_name]
        start_column = ws.max_column + 1 if ws.max_column else 1
    else:
        ws = wb.create_sheet(fund_name)

    for col_index, col_name in enumerate(combined_df_temp.columns, start=start_column):
        ws.cell(row=1, column=col_index)  # Write column names
        for row_index, data in enumerate(combined_df_temp[col_name], start=2):
            ws.cell(row=row_index, column=col_index, value=data)  # Write data

    wb.save(excel_filename)
    print(f"Data added to '{fund_name}' sheet in {excel_filename}")

driver.quit()


###############################################################################################################
