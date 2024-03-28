from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import re

wait_duration = 20

urls = [
    "https://www.moneycontrol.com/mutual-funds/icici-prudential-value-discovery-fund/investment-info/MPI087",
    "https://www.moneycontrol.com/mutual-funds/sbi-contra-fund/investment-info/MSB010",
    "https://www.moneycontrol.com/mutual-funds/quant-small-cap-fund-direct-plan/investment-info/MES056",
    "https://www.moneycontrol.com/mutual-funds/axis-small-cap-fund-direct-plan/investment-info/MAA316",
    "https://www.moneycontrol.com/mutual-funds/kotak-small-cap-fund-direct-plan/investment-info/MKM516",
    "https://www.moneycontrol.com/mutual-funds/nippon-india-small-cap-fund/investment-info/MRC587",
    "https://www.moneycontrol.com/mutual-funds/quant-active-fund-direct-plan/investment-info/MES061",
    "https://www.moneycontrol.com/mutual-funds/kotak-emerging-equity-scheme-regular-plan/investment-info/MKM099",
    "https://www.moneycontrol.com/mutual-funds/tata-large-cap-fund-regular-plan/investment-info/MTA006",
    "https://www.moneycontrol.com/mutual-funds/sbi-blue-chip-fund/investment-info/MSB079",
    "https://www.moneycontrol.com/mutual-funds/nav/axis-bluechip-fund-growth/MAA009",
    "https://www.moneycontrol.com/mutual-funds/nippon-india-large-cap-fund-regular-plan/investment-info/MRC155",
    "https://www.moneycontrol.com/mutual-funds/quant-large-mid-cap-fund/investment-info/MES022",
    "https://www.moneycontrol.com/mutual-funds/tata-large-mid-cap-fund-regular-plan/investment-info/MTA043",
    "https://www.moneycontrol.com/mutual-funds/canara-robeco-blue-chip-equity-fund-regular-plan/investment-info/MCA174",
    "https://www.moneycontrol.com/mutual-funds/mirae-asset-emerging-bluechip-fund/investment-info/MMA088",
    "https://www.moneycontrol.com/mutual-funds/edelweiss-mid-cap-fund-direct-plan/investment-info/MJP117",
    "https://www.moneycontrol.com/mutual-funds/parag-parikh-flexi-cap-fund-regular-plan/investment-info/MPP001",
    "https://www.moneycontrol.com/mutual-funds/icici-prudential-india-opportunities-fund-regular-plan/investment-info/MPI4087",
    "https://www.moneycontrol.com/mutual-funds/nippon-india-flexi-cap-fund-direct-plan/investment-info/MRC2885",
    "https://www.moneycontrol.com/mutual-funds/quant-large-cap-fund-direct-plan/investment-info/MES080",
    "https://www.moneycontrol.com/mutual-funds/bnp-paribas-long-term-equity-fund/investment-info/MAB041",
    "https://www.moneycontrol.com/mutual-funds/icici-prudential-multi-asset-fund/investment-info/MPI038",
    "https://www.moneycontrol.com/mutual-funds/icici-prudential-technology-fund/investment-info/MPI015",
    "https://www.moneycontrol.com/mutual-funds/icici-prudential-value-discovery-fund/investment-info/MPI087",
    "https://www.moneycontrol.com/mutual-funds/sbi-contra-fund/investment-info/MSB010",
    "https://www.moneycontrol.com/mutual-funds/quant-small-cap-fund-direct-plan/investment-info/MES056",
    "https://www.moneycontrol.com/mutual-funds/axis-small-cap-fund-direct-plan/investment-info/MAA286",
    "https://www.moneycontrol.com/mutual-funds/quant-active-fund-direct-plan/investment-info/MQU001",
    "https://www.moneycontrol.com/mutual-funds/axis-mid-cap-fund-direct-plan/investment-info/MAA273",
    "https://www.moneycontrol.com/mutual-funds/sbi-midcap-fund-direct-plan/investment-info/MSB260",
    "https://www.moneycontrol.com/mutual-funds/tata-mid-cap-growth-fund-direct-plan/investment-info/MTA210",
    "https://www.moneycontrol.com/mutual-funds/nippon-india-small-cap-fund-direct-plan/investment-info/MRC935",
    "https://www.moneycontrol.com/mutual-funds/nav/quant-mid-cap-fund-direct/MES043",
    "https://www.moneycontrol.com/mutual-funds/nippon-india-growth-fund-direct-plan/investment-info/MRC919",
]

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)
try:
    for index, url in enumerate(urls):
        try:
            driver.get(url)

            # Extract fund name from the webpage URL
            fund_name = re.search(r"mutual-funds\/(.*?)\/investment-info", url).group(1)
            fund_name = fund_name.replace("-", " ").title()  # Format fund name

            parent_div_xpath = "/html/body/div[15]/section[3]/div/div/div[1]"

            h2_elements = driver.find_elements(By.XPATH, f"{parent_div_xpath}//h2")
            combined_data_h2 = [elem.text for elem in h2_elements]

            p_elements = driver.find_elements(By.XPATH, f"{parent_div_xpath}//p")
            combined_data_p = [elem.text for elem in p_elements]

            span_elements = driver.find_elements(By.XPATH, f"{parent_div_xpath}//p")
            combined_data_span = [elem.text for elem in span_elements]

            print("Data extracted successfully (H2):", combined_data_h2)
            print("Data extracted successfully (P):", combined_data_p)

            excel_filename = "fund_objective.xlsx"
            wb = load_workbook(excel_filename)

            ws = wb.create_sheet(title=fund_name)

            start_column = ws.max_column + 1 if ws.max_column else 1

            for row, value in enumerate(combined_data_h2, start=1):
                ws.cell(row=row, column=start_column, value=value)
            for row, value in enumerate(combined_data_p, start=1):
                ws.cell(row=row, column=start_column, value=value)

            wb.save(excel_filename)
            print(f"Data added to '{fund_name}' sheet in {excel_filename}")

        except Exception as e:
            print(urls)
            print(f"Error processing URL: {url}. Error: {str(e)}")

finally:
    driver.quit()
