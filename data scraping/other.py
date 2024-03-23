from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from urllib.parse import urlparse
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

wait_duration = 5

urls = [
    "https://www.valueresearchonline.com/funds/2310/icici-prudential-value-discovery-fund/",
    "https://www.valueresearchonline.com/funds/633/sbi-contra-fund/",
    "https://www.valueresearchonline.com/funds/66/quant-small-cap-fund/",
    "https://www.valueresearchonline.com/funds/22334/axis-small-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/2597/kotak-small-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/11463/nippon-india-small-cap-fund/",
    "https://www.valueresearchonline.com/funds/952/quant-active-fund/",
    "https://www.valueresearchonline.com/funds/958/quant-mid-cap-fund/",
    "https://www.valueresearchonline.com/funds/12052/axis-midcap-fund/",
    "https://www.valueresearchonline.com/funds/4270/kotak-emerging-equity-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/183/nippon-india-growth-fund/",
    "https://www.valueresearchonline.com/funds/2662/sbi-magnum-midcap-fund/",
    "https://www.valueresearchonline.com/funds/103/tata-midcap-growth-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/509/tata-large-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/4871/edelweiss-large-mid-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/3083/sbi-bluechip-fund/",
    "https://www.valueresearchonline.com/funds/10780/axis-bluechip-fund/",
    "https://www.valueresearchonline.com/funds/5270/nippon-india-large-cap-fund/",
    "https://www.valueresearchonline.com/funds/577/kotak-bluechip-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/4111/quant-large-and-mid-cap-fund/",
    "https://www.valueresearchonline.com/funds/102/tata-large-mid-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/10432/edelweiss-large-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/197/sbi-large-midcap-fund/",
    "https://www.valueresearchonline.com/funds/11333/canara-robeco-bluechip-equity-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/11213/mirae-asset-emerging-bluechip-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/6227/edelweiss-mid-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/19699/parag-parikh-flexi-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/38583/icici-prudential-india-opportunities-fund/",
    "https://www.valueresearchonline.com/funds/41614/nippon-india-flexi-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/42365/quant-large-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/3001/baroda-bnp-paribas-elss-fund/",
    "https://www.valueresearchonline.com/funds/8216/quant-focused-fund/",
    "https://www.valueresearchonline.com/funds/40104/parag-parikh-tax-saver-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/10032/bandhan-elss-tax-saver-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/2937/kotak-elss-tax-saver-regular-plan/",
    "https://www.valueresearchonline.com/funds/42502/sundaram-flexi-cap-fund-regular-plan/",
    "https://www.valueresearchonline.com/funds/1471/icici-prudential-multi-asset-fund/",
    "https://www.valueresearchonline.com/funds/737/icici-prudential-technology-fund/",
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#other"

for index, url in enumerate(urls):
    try:
        chrome_service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=chrome_service)

        driver.get(url)

        parent_div_xpath_h3 = (
            "/html/body/section[2]/div[2]/div/div[7]/section/div/div[1]"
        )

        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_h3))
        )

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        parsed_url = urlparse(url)
        fund_name = parsed_url.path.split("/")[3].replace("-", " ").title()[:31]

        excel_filename = "combined_data.xlsx"
        wb = load_workbook(excel_filename)

        if fund_name in wb.sheetnames:
            ws = wb[fund_name]
            start_column = ws.max_column + 1 if ws.max_column else 1
        else:
            # Create a new sheet with the fund name
            ws = wb.create_sheet(fund_name)
            start_column = 1

        for i, data in enumerate(combined_data_li, start=1):
            ws.cell(row=i, column=start_column, value=data)

        wb.save(excel_filename)
        print(f"Data added to '{fund_name}' sheet in {excel_filename}")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")
driver.quit()
