from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import re

wait_duration = 20

urls = ["https://groww.in/mutual-funds/quant-small-cap-fund-direct-plan-growth"]

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)
try:
    for index, url in enumerate(urls):
        try:
            driver.get(url)

            # # Extract fund name from the webpage URL
            # fund_name = re.search(r"mutual-funds\/(.*?)\/investment-info", url).group(1)
            # fund_name = fund_name.replace("-", " ").title()  # Format fund name
            fund_name = "Quant"

            parent_div_xpath = "/html/body"

            h2_elements = driver.find_elements(By.XPATH, f"{parent_div_xpath}//div")
            combined_data_h2 = [elem.text for elem in h2_elements]

            # p_elements = driver.find_elements(By.XPATH, f"{parent_div_xpath}//p")
            # combined_data_p = [elem.text for elem in p_elements]

            # span_elements = driver.find_elements(By.XPATH, f"{parent_div_xpath}//p")
            # combined_data_span = [elem.text for elem in span_elements]

            print("Data extracted successfully (H2):", combined_data_h2)
            # print("Data extracted successfully (P):", combined_data_p)

            excel_filename = "combined_data.xlsx"
            wb = load_workbook(excel_filename)

            ws = wb.create_sheet(title=fund_name)

            start_column = ws.max_column + 1 if ws.max_column else 1

            for row, value in enumerate(combined_data_h2, start=1):
                ws.cell(row=row, column=start_column, value=value)
            # for row, value in enumerate(combined_data_p, start=1):
            #     ws.cell(row=row, column=start_column, value=value)

            wb.save(excel_filename)
            print(f"Data added to '{fund_name}' sheet in {excel_filename}")

        except Exception as e:
            print(f"Error processing URL: {url}. Error: {str(e)}")

finally:
    driver.quit()
