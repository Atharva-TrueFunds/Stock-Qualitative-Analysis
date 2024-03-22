from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from urllib.parse import urlparse
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time


# Define wait_duration
wait_duration = 5  # 5 seconds, adjust as needed

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)


# Function to process tables
def process_tables(url):
    driver.get(url)
    tables = driver.find_elements(By.TAG_NAME, "table")
    dfs = []
    for table in tables:
        table_html = table.get_attribute("outerHTML")
        df = pd.read_html(table_html)
        if df:
            df = df[0].dropna(how="all").reset_index(drop=True)
            dfs.append(df)
    return dfs


urls = [
    "https://www.valueresearchonline.com/funds/633/sbi-contra-fund/",
    # Add more URLs as needed
]

for i in range(len(urls)):
    urls[i] = urls[i] + "#fund-portfolio"

for url in urls:
    try:
        driver.get(url)

        # Define XPaths
        parent_div_xpath_p = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[3]"
        parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[2]/div/div[1]/div/div/div[1]/div[1]"
        parent_div_xpath_li = (
            "/html/body/section[2]/div[2]/div/div[5]/div/div/section[2]/div[1]/div/ul"
        )

        # Wait for visibility of elements
        WebDriverWait(driver, wait_duration).until(
            EC.visibility_of_element_located((By.XPATH, parent_div_xpath_p))
        )

        parent_div_p = driver.find_element(By.XPATH, parent_div_xpath_p)
        list_items_p = parent_div_p.find_elements(By.XPATH, ".//p")
        combined_data_p = [item.text for item in list_items_p]

        parent_div_h3 = driver.find_element(By.XPATH, parent_div_xpath_h3)
        list_items_h3 = parent_div_h3.find_elements(By.XPATH, ".//p")
        combined_data_h3 = [h3.text for h3 in list_items_h3]

        parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_li)
        list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
        combined_data_li = [item.text for item in list_items_li]

        # Check if all arrays have the same length
        if len(combined_data_h3) == len(combined_data_p) == len(combined_data_li):
            # Combine all data into a DataFrame
            combined_df_temp = pd.DataFrame(
                {
                    "List Items (h3)": combined_data_h3,
                    "List Items (p)": combined_data_p,
                    "List Items (li)": combined_data_li,
                }
            )

            # Get fund name from URL
            parsed_url = urlparse(url)
            fund_name = parsed_url.path.split("/")[3]

            # Load workbook
            excel_filename = "combined_data.xlsx"
            wb = load_workbook(excel_filename)

            # Check if sheet exists
            if fund_name in wb.sheetnames:
                ws = wb[fund_name]  # Get the existing sheet
                start_column = (
                    ws.max_column + 1 if ws.max_column else 1
                )  # Get the next available column
            else:
                ws = wb.create_sheet(fund_name)  # Create a new sheet with fund name
                start_column = 1  # Start from the first column

            # Add data to the sheet
            for i, data in enumerate(combined_data_li, start=1):
                ws.cell(row=i, column=start_column, value=data)

            # Save Excel file
            wb.save(excel_filename)
            print(f"Data added to '{fund_name}' sheet in {excel_filename}")
        else:
            print("Arrays have different lengths. Cannot create DataFrame.")

    except Exception as e:
        print(f"Error processing URL: {url}. Error: {str(e)}")

driver.quit()
