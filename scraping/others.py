from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from urllib.parse import urlparse
from openpyxl import load_workbook

url = "https://www.valueresearchonline.com/funds/66/quant-small-cap-fund/#other"

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)
driver.get(url)

parent_div_xpath_h3 = "/html/body/section[2]/div[2]/div/div[7]/section/div/div[1]"
parent_div_li = driver.find_element(By.XPATH, parent_div_xpath_h3)
list_items_li = parent_div_li.find_elements(By.XPATH, ".//span")
combined_data_li = [item.text for item in list_items_li]

parsed_url = urlparse(url)
fund_name = parsed_url.path.split("/")[3]

excel_filename = "combined_data.xlsx"
wb = load_workbook(excel_filename)

if fund_name in wb.sheetnames:
    ws = wb[fund_name]
    start_column = ws.max_column + 1 if ws.max_column else 1
else:
    # Create a new sheet with the fund name
    ws = wb.create_sheet(fund_name)
    start_column = 1

for i, data in enumerate(combined_data_li, start=1):
    ws.cell(row=i, column=start_column, value=data)

wb.save(excel_filename)
print(f"Data added to '{fund_name}' sheet in {excel_filename}")

driver.quit()
