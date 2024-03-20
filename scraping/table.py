from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import pandas as pd
from openpyxl import load_workbook

chrome_service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=chrome_service)
url = "https://www.valueresearchonline.com/funds/17366/quant-small-cap-fund-direct-plan/#performance"
driver.get(url)

# Process all tables on the webpage
tables = driver.find_elements(By.TAG_NAME, "table")
all_dfs = []
for table in tables:
    table_html = table.get_attribute("outerHTML")
    df = pd.read_html(table_html)
    if df:
        df = df[0].dropna(how="all").reset_index(drop=True)
        all_dfs.append(df)

# Concatenate all DataFrames
if all_dfs:
    combined_df = pd.concat(all_dfs, ignore_index=True)

    # Clean up column names (remove tuples)
    combined_df.columns = [
        col[0] if isinstance(col, tuple) else col for col in combined_df.columns
    ]

    # Save to Excel
    excel_filename = "combined_data.xlsx"
    wb = load_workbook(excel_filename)
    ws = wb.active

    # Find the first empty column
    start_column = ws.max_column + 1

    # Write column names to the worksheet
    for col_index, column in enumerate(combined_df.columns, start=start_column):
        ws.cell(row=1, column=col_index, value=column)

    # Write data to the worksheet
    start_row = 2
    for _, row in combined_df.iterrows():
        for col_index, value in enumerate(row, start=start_column):
            ws.cell(row=start_row, column=col_index, value=value)
        start_row += 1

    wb.save(excel_filename)
    print(f"All DataFrames saved to {excel_filename}")
else:
    print("No tables found on the webpage.")

driver.quit()
