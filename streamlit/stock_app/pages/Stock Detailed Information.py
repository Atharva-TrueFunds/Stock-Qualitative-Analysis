import streamlit as st
import pandas as pd
from bsedata.bse import BSE
from datetime import datetime
import os

today = datetime.today().strftime("%d %b")
sheet_name = today


def fetch_data_from_bse(scripCode, sheet_name):
    bse = BSE()
    stock_data = []

    try:
        for code in scripCode:
            try:
                stock_quote = bse.getQuote(str(code))
                fetched_data = pd.DataFrame(stock_quote)
                fetched_data["updatedOn"] = pd.to_datetime(
                    fetched_data["updatedOn"], format="%d %b %y | %I:%M %p"
                )

                fetched_data["date_part"] = fetched_data["updatedOn"].dt.strftime(
                    "%d %b"
                )

                if fetched_data["date_part"].iloc[0] != sheet_name:
                    na_data = {
                        "scripCode": code,
                        "companyName": fetched_data["companyName"].iloc[0],
                        "currentValue": "NA",
                        "change": "NA",
                        "pChange": "NA",
                        "updatedOn": "NA",
                        "securityID": "NA",
                        "group": "NA",
                        "faceValue": "NA",
                        "industry": "NA",
                        "previousClose": "NA",
                        "previousOpen": "NA",
                        "dayHigh": "NA",
                        "dayLow": "NA",
                        "52weekHigh": "NA",
                        "52weekLow": "NA",
                        "weightedAvgPrice": "NA",
                        "totalTradedValue": "NA",
                        "totalTradedQuantity": "NA",
                        "2WeekAvgQuantity": "NA",
                        "marketCapFull": "NA",
                        "marketCapFreeFloat": "NA",
                        "buy": "NA",
                        "sell": "NA",
                    }
                    na_df = pd.DataFrame(na_data, index=[0])
                    stock_data.append(na_df)
                else:
                    fetched_data = fetched_data.drop(columns=["date_part"])
                    stock_data.append(fetched_data.head(1))
            except Exception as e:
                st.error(f"Error fetching data for script code {code}: {e}")
                na_data = {
                    "scripCode": code,
                    "companyName": "NA",
                    "currentValue": "NA",
                    "change": "NA",
                    "pChange": "NA",
                    "updatedOn": "NA",
                    "securityID": "NA",
                    "group": "NA",
                    "faceValue": "NA",
                    "industry": "NA",
                    "previousClose": "NA",
                    "previousOpen": "NA",
                    "dayHigh": "NA",
                    "dayLow": "NA",
                    "52weekHigh": "NA",
                    "52weekLow": "NA",
                    "weightedAvgPrice": "NA",
                    "totalTradedValue": "NA",
                    "totalTradedQuantity": "NA",
                    "2WeekAvgQuantity": "NA",
                    "marketCapFull": "NA",
                    "marketCapFreeFloat": "NA",
                    "buy": "NA",
                    "sell": "NA",
                }
                na_df = pd.DataFrame(na_data, index=[0])
                stock_data.append(na_df)

        if stock_data:
            combined_data = pd.concat(stock_data, ignore_index=True)
            print(combined_data.columns)
            return combined_data
        else:
            return pd.DataFrame()

    except Exception as e:
        st.error(f"Error: {e}")


def main():
    st.title("BSE Stock Information")
    run_button_clicked = st.button("Run")

    if run_button_clicked:
        try:
            excel_data = pd.read_excel(
                r"C:\Users\HP\Staging\streamlit\stock_app\code.xlsx",
                engine="openpyxl",
            )
            excel_data_1 = pd.read_excel(
                r"C:\Users\HP\Staging\streamlit\stock_app\code.xlsx"
            )
            selected_columns = excel_data_1[
                ["Sector_Name", "Industry_New_Name", "Igroup_Name", "Isubgroup_Name"]
            ]
            scripCode = excel_data["scripCode"].tolist()
            fetched_data = fetch_data_from_bse(scripCode, sheet_name)
            concatenated_data = pd.concat([fetched_data, selected_columns], axis=1)
            st.write(concatenated_data)
            st.write(f"Saving fetched data to Excel, new sheet name: {sheet_name}...")

            with pd.ExcelWriter(
                "stock_quote.xlsx", mode="a", engine="openpyxl"
            ) as writer:

                concatenated_data.to_excel(writer, sheet_name=sheet_name, index=False)

            st.success("Data saved to stock_quote.xlsx successfully!")
        except Exception as e:
            st.error(f"Error: {e}")

        ##########    currentValue

        excel_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\stock_quote.xlsx"
        current_value_file_path = (
            r"C:\Users\HP\Staging\streamlit\stock_app\currentValue.xlsx"
        )
        df_dict = pd.read_excel(excel_file_path, sheet_name=sheet_name)
        df_current = pd.read_excel(current_value_file_path)
        try:
            symbols = df_dict["scripCode"].tolist()
            b = BSE()
            stock_data = []

            for symbol in symbols:
                symbol_data = df_dict[df_dict["scripCode"] == symbol]

                if not symbol_data.empty:
                    stock_data.append(
                        {
                            "scripCode": symbol,
                            "companyName": symbol_data["companyName"].values[0],
                            sheet_name: symbol_data["currentValue"].values[0],
                        }
                    )
                else:
                    print(f"No data found for symbol: {symbol}")
            df_stock = pd.DataFrame(stock_data)
            with pd.ExcelWriter(
                current_value_file_path, mode="a", engine="openpyxl"
            ) as writer:
                df_stock.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            st.error(f"Error: {e}")
        from openpyxl import load_workbook

        try:
            workbook = load_workbook("currentValue.xlsx")
            source_sheet = workbook[sheet_name]
            workbook_all_sheets = load_workbook("All_sheets.xlsx")
            destination_sheet = workbook_all_sheets["currentValue"]
            last_column_index = source_sheet.max_column
            last_column_values = []
            for row in source_sheet.iter_rows(
                min_row=1,
                max_row=source_sheet.max_row,
                min_col=last_column_index,
                max_col=last_column_index,
                values_only=True,
            ):
                for value in row:
                    last_column_values.append(value)
            next_column_index = destination_sheet.max_column + 1
            for i, value in enumerate(last_column_values, start=1):
                destination_sheet.cell(row=i, column=next_column_index, value=value)
            workbook_all_sheets.save("All_sheets.xlsx")
        except Exception as e:
            print(f"Error: {e}")

        # ##########    52weekHigh

        excel_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\stock_quote.xlsx"
        weekHigh_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\weekHigh52.xlsx"
        df_dict = pd.read_excel(excel_file_path, sheet_name=sheet_name)
        df_current = pd.read_excel(weekHigh_file_path)
        try:
            symbols = df_dict["scripCode"].tolist()
            b = BSE()
            stock_data = []

            for symbol in symbols:
                symbol_data = df_dict[df_dict["scripCode"] == symbol]

                if not symbol_data.empty:
                    stock_data.append(
                        {
                            "scripCode": symbol,
                            "companyName": symbol_data["companyName"].values[0],
                            sheet_name: symbol_data["52weekHigh"].values[0],
                        }
                    )
                else:
                    print(f"No data found for symbol: {symbol}")
            df_stock = pd.DataFrame(stock_data)
            with pd.ExcelWriter(
                weekHigh_file_path, mode="a", engine="openpyxl"
            ) as writer:
                df_stock.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            st.error(f"Error: {e}")
        from openpyxl import load_workbook

        try:
            workbook = load_workbook("weekHigh52.xlsx")
            source_sheet = workbook[sheet_name]
            workbook_all_sheets = load_workbook("All_sheets.xlsx")
            destination_sheet = workbook_all_sheets["52weekHigh"]
            last_column_index = source_sheet.max_column
            last_column_values = []
            for row in source_sheet.iter_rows(
                min_row=1,
                max_row=source_sheet.max_row,
                min_col=last_column_index,
                max_col=last_column_index,
                values_only=True,
            ):
                for value in row:
                    last_column_values.append(value)
            next_column_index = destination_sheet.max_column + 1
            for i, value in enumerate(last_column_values, start=1):
                destination_sheet.cell(row=i, column=next_column_index, value=value)
            workbook_all_sheets.save("All_sheets.xlsx")
        except Exception as e:
            print(f"Error: {e}")

        # ##########    52weekLow

        excel_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\stock_quote.xlsx"
        weekLow_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\weekLow52.xlsx"
        df_dict = pd.read_excel(excel_file_path, sheet_name=sheet_name)
        df_current = pd.read_excel(weekLow_file_path)
        try:
            symbols = df_dict["scripCode"].tolist()
            b = BSE()
            stock_data = []

            for symbol in symbols:
                symbol_data = df_dict[df_dict["scripCode"] == symbol]

                if not symbol_data.empty:
                    stock_data.append(
                        {
                            "scripCode": symbol,
                            "companyName": symbol_data["companyName"].values[0],
                            sheet_name: symbol_data["52weekLow"].values[0],
                        }
                    )
                else:
                    print(f"No data found for symbol: {symbol}")
            df_stock = pd.DataFrame(stock_data)
            # st.write(df_stock)
            # st.write("Saving fetched data to Excel, new sheet name: sheet_name...")
            with pd.ExcelWriter(
                weekLow_file_path, mode="a", engine="openpyxl"
            ) as writer:
                df_stock.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            st.error(f"Error: {e}")
        from openpyxl import load_workbook

        try:
            workbook = load_workbook("weekLow52.xlsx")
            source_sheet = workbook[sheet_name]
            workbook_all_sheets = load_workbook("All_sheets.xlsx")
            destination_sheet = workbook_all_sheets["52weekLow"]
            last_column_index = source_sheet.max_column
            last_column_values = []
            for row in source_sheet.iter_rows(
                min_row=1,
                max_row=source_sheet.max_row,
                min_col=last_column_index,
                max_col=last_column_index,
                values_only=True,
            ):
                for value in row:
                    last_column_values.append(value)
            next_column_index = destination_sheet.max_column + 1
            for i, value in enumerate(last_column_values, start=1):
                destination_sheet.cell(row=i, column=next_column_index, value=value)
            workbook_all_sheets.save("All_sheets.xlsx")
        except Exception as e:
            print(f"Error: {e}")

        # ##########    change

        excel_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\stock_quote.xlsx"
        change_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\change.xlsx"
        df_dict = pd.read_excel(excel_file_path, sheet_name=sheet_name)
        df_current = pd.read_excel(change_file_path)
        try:
            symbols = df_dict["scripCode"].tolist()
            b = BSE()
            stock_data = []

            for symbol in symbols:
                symbol_data = df_dict[df_dict["scripCode"] == symbol]

                if not symbol_data.empty:
                    stock_data.append(
                        {
                            "scripCode": symbol,
                            "companyName": symbol_data["companyName"].values[0],
                            sheet_name: symbol_data["change"].values[0],
                        }
                    )
                else:
                    print(f"No data found for symbol: {symbol}")
            df_stock = pd.DataFrame(stock_data)
            # st.write(df_stock)
            # st.write("Saving fetched data to Excel, new sheet name: sheet_name...")
            with pd.ExcelWriter(
                change_file_path, mode="a", engine="openpyxl"
            ) as writer:
                df_stock.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            st.error(f"Error: {e}")
        from openpyxl import load_workbook

        try:
            workbook = load_workbook("change.xlsx")
            source_sheet = workbook[sheet_name]
            workbook_all_sheets = load_workbook("All_sheets.xlsx")
            destination_sheet = workbook_all_sheets["Change"]
            last_column_index = source_sheet.max_column
            last_column_values = []
            for row in source_sheet.iter_rows(
                min_row=1,
                max_row=source_sheet.max_row,
                min_col=last_column_index,
                max_col=last_column_index,
                values_only=True,
            ):
                for value in row:
                    last_column_values.append(value)
            next_column_index = destination_sheet.max_column + 1
            for i, value in enumerate(last_column_values, start=1):
                destination_sheet.cell(row=i, column=next_column_index, value=value)
            workbook_all_sheets.save("All_sheets.xlsx")
        except Exception as e:
            print(f"Error: {e}")

        # ########## pChange
        excel_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\stock_quote.xlsx"
        pChange_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\pChange.xlsx"
        df_dict = pd.read_excel(excel_file_path, sheet_name=sheet_name)
        df_current = pd.read_excel(pChange_file_path)
        try:
            symbols = df_dict["scripCode"].tolist()
            b = BSE()
            stock_data = []

            for symbol in symbols:
                symbol_data = df_dict[df_dict["scripCode"] == symbol]

                if not symbol_data.empty:
                    stock_data.append(
                        {
                            "scripCode": symbol,
                            "companyName": symbol_data["companyName"].values[0],
                            sheet_name: symbol_data["pChange"].values[0],
                        }
                    )
                else:
                    print(f"No data found for symbol: {symbol}")
            df_stock = pd.DataFrame(stock_data)
            # st.write(df_stock)
            # st.write("Saving fetched data to Excel, new sheet name: sheet_name...")
            with pd.ExcelWriter(
                pChange_file_path, mode="a", engine="openpyxl"
            ) as writer:
                df_stock.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            st.error(f"Error: {e}")
        from openpyxl import load_workbook

        try:
            workbook = load_workbook("pChange.xlsx")
            source_sheet = workbook[sheet_name]
            workbook_all_sheets = load_workbook("All_sheets.xlsx")
            destination_sheet = workbook_all_sheets["pChange"]
            last_column_index = source_sheet.max_column
            last_column_values = []
            for row in source_sheet.iter_rows(
                min_row=1,
                max_row=source_sheet.max_row,
                min_col=last_column_index,
                max_col=last_column_index,
                values_only=True,
            ):
                for value in row:
                    last_column_values.append(value)
            next_column_index = destination_sheet.max_column + 1
            for i, value in enumerate(last_column_values, start=1):
                destination_sheet.cell(row=i, column=next_column_index, value=value)
            workbook_all_sheets.save("All_sheets.xlsx")
        except Exception as e:
            print(f"Error: {e}")

        # # previousClose
        excel_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\stock_quote.xlsx"
        previousClose_file_path = (
            r"C:\Users\HP\Staging\streamlit\stock_app\previousClose.xlsx"
        )
        df_dict = pd.read_excel(excel_file_path, sheet_name=sheet_name)
        df_current = pd.read_excel(previousClose_file_path)
        try:
            symbols = df_dict["scripCode"].tolist()
            b = BSE()
            stock_data = []

            for symbol in symbols:
                symbol_data = df_dict[df_dict["scripCode"] == symbol]

                if not symbol_data.empty:
                    stock_data.append(
                        {
                            "scripCode": symbol,
                            "companyName": symbol_data["companyName"].values[0],
                            sheet_name: symbol_data["previousClose"].values[0],
                        }
                    )
                else:
                    print(f"No data found for symbol: {symbol}")
            df_stock = pd.DataFrame(stock_data)
            # st.write(df_stock)
            # st.write("Saving fetched data to Excel, new sheet name: sheet_name...")
            with pd.ExcelWriter(
                previousClose_file_path, mode="a", engine="openpyxl"
            ) as writer:
                df_stock.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            st.error(f"Error: {e}")
        from openpyxl import load_workbook

        try:
            workbook = load_workbook("previousClose.xlsx")
            source_sheet = workbook[sheet_name]
            workbook_all_sheets = load_workbook("All_sheets.xlsx")
            destination_sheet = workbook_all_sheets["previousClose"]
            last_column_index = source_sheet.max_column
            last_column_values = []
            for row in source_sheet.iter_rows(
                min_row=1,
                max_row=source_sheet.max_row,
                min_col=last_column_index,
                max_col=last_column_index,
                values_only=True,
            ):
                for value in row:
                    last_column_values.append(value)
            next_column_index = destination_sheet.max_column + 1
            for i, value in enumerate(last_column_values, start=1):
                destination_sheet.cell(row=i, column=next_column_index, value=value)
            workbook_all_sheets.save("All_sheets.xlsx")
        except Exception as e:
            print(f"Error: {e}")

            excel_file_path = (
                r"C:\Users\HP\Staging\streamlit\stock_app\stock_quote.xlsx"
            )

        # # previousOpen
        excel_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\stock_quote.xlsx"
        previousOpen_file_path = (
            r"C:\Users\HP\Staging\streamlit\stock_app\previousOpen.xlsx"
        )
        df_dict = pd.read_excel(excel_file_path, sheet_name=sheet_name)
        df_current = pd.read_excel(previousOpen_file_path)
        try:
            symbols = df_dict["scripCode"].tolist()
            b = BSE()
            stock_data = []

            for symbol in symbols:
                symbol_data = df_dict[df_dict["scripCode"] == symbol]

                if not symbol_data.empty:
                    stock_data.append(
                        {
                            "scripCode": symbol,
                            "companyName": symbol_data["companyName"].values[0],
                            sheet_name: symbol_data["previousOpen"].values[0],
                        }
                    )
                else:
                    print(f"No data found for symbol: {symbol}")
            df_stock = pd.DataFrame(stock_data)
            # st.write(df_stock)
            # st.write("Saving fetched data to Excel, new sheet name: sheet_name...")
            with pd.ExcelWriter(
                previousOpen_file_path, mode="a", engine="openpyxl"
            ) as writer:
                df_stock.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            st.error(f"Error: {e}")
        from openpyxl import load_workbook

        try:
            workbook = load_workbook("previousOpen.xlsx")
            source_sheet = workbook[sheet_name]
            workbook_all_sheets = load_workbook("All_sheets.xlsx")
            destination_sheet = workbook_all_sheets["previousOpen"]
            last_column_index = source_sheet.max_column
            last_column_values = []
            for row in source_sheet.iter_rows(
                min_row=1,
                max_row=source_sheet.max_row,
                min_col=last_column_index,
                max_col=last_column_index,
                values_only=True,
            ):
                for value in row:
                    last_column_values.append(value)
            next_column_index = destination_sheet.max_column + 1
            for i, value in enumerate(last_column_values, start=1):
                destination_sheet.cell(row=i, column=next_column_index, value=value)
            workbook_all_sheets.save("All_sheets.xlsx")
        except Exception as e:
            print(f"Error: {e}")

        # # dayHigh
        excel_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\stock_quote.xlsx"
        dayHigh_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\dayHigh.xlsx"
        df_dict = pd.read_excel(excel_file_path, sheet_name=sheet_name)
        df_current = pd.read_excel(dayHigh_file_path)
        try:
            symbols = df_dict["scripCode"].tolist()
            b = BSE()
            stock_data = []

            for symbol in symbols:
                symbol_data = df_dict[df_dict["scripCode"] == symbol]

                if not symbol_data.empty:
                    stock_data.append(
                        {
                            "scripCode": symbol,
                            "companyName": symbol_data["companyName"].values[0],
                            sheet_name: symbol_data["dayHigh"].values[0],
                        }
                    )
                else:
                    print(f"No data found for symbol: {symbol}")
            df_stock = pd.DataFrame(stock_data)
            # st.write(df_stock)
            # st.write("Saving fetched data to Excel, new sheet name: sheet_name...")
            with pd.ExcelWriter(
                dayHigh_file_path, mode="a", engine="openpyxl"
            ) as writer:
                df_stock.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            st.error(f"Error: {e}")
        from openpyxl import load_workbook

        try:
            workbook = load_workbook("dayHigh.xlsx")
            source_sheet = workbook[sheet_name]
            workbook_all_sheets = load_workbook("All_sheets.xlsx")
            destination_sheet = workbook_all_sheets["dayHigh"]
            last_column_index = source_sheet.max_column
            last_column_values = []
            for row in source_sheet.iter_rows(
                min_row=1,
                max_row=source_sheet.max_row,
                min_col=last_column_index,
                max_col=last_column_index,
                values_only=True,
            ):
                for value in row:
                    last_column_values.append(value)
            next_column_index = destination_sheet.max_column + 1
            for i, value in enumerate(last_column_values, start=1):
                destination_sheet.cell(row=i, column=next_column_index, value=value)
            workbook_all_sheets.save("ALl_sheets.xlsx")
        except Exception as e:
            print(f"Error: {e}")

        # # dayLow
        excel_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\stock_quote.xlsx"
        dayLow_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\dayLow.xlsx"
        df_dict = pd.read_excel(excel_file_path, sheet_name=sheet_name)
        df_current = pd.read_excel(dayLow_file_path)
        try:
            symbols = df_dict["scripCode"].tolist()
            b = BSE()
            stock_data = []

            for symbol in symbols:
                symbol_data = df_dict[df_dict["scripCode"] == symbol]

                if not symbol_data.empty:
                    stock_data.append(
                        {
                            "scripCode": symbol,
                            "companyName": symbol_data["companyName"].values[0],
                            sheet_name: symbol_data["dayLow"].values[0],
                        }
                    )
                else:
                    print(f"No data found for symbol: {symbol}")
            df_stock = pd.DataFrame(stock_data)
            # st.write(df_stock)
            # st.write("Saving fetched data to Excel, new sheet name: sheet_name...")
            with pd.ExcelWriter(
                dayLow_file_path, mode="a", engine="openpyxl"
            ) as writer:
                df_stock.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            st.error(f"Error: {e}")
        from openpyxl import load_workbook

        try:
            workbook = load_workbook("dayLow.xlsx")
            source_sheet = workbook[sheet_name]
            workbook_all_sheets = load_workbook("All_sheets.xlsx")
            destination_sheet = workbook_all_sheets["dayLow"]
            last_column_index = source_sheet.max_column
            last_column_values = []
            for row in source_sheet.iter_rows(
                min_row=1,
                max_row=source_sheet.max_row,
                min_col=last_column_index,
                max_col=last_column_index,
                values_only=True,
            ):
                for value in row:
                    last_column_values.append(value)
            next_column_index = destination_sheet.max_column + 1
            for i, value in enumerate(last_column_values, start=1):
                destination_sheet.cell(row=i, column=next_column_index, value=value)
            workbook_all_sheets.save("All_sheets.xlsx")
        except Exception as e:
            print(f"Error: {e}")

        # # weightedAvgPrice
        excel_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\stock_quote.xlsx"
        weightedAvgPrice_file_path = (
            r"C:\Users\HP\Staging\streamlit\stock_app\weightedAvgPrice.xlsx"
        )
        df_dict = pd.read_excel(excel_file_path, sheet_name=sheet_name)
        df_current = pd.read_excel(weightedAvgPrice_file_path)
        try:
            symbols = df_dict["scripCode"].tolist()
            b = BSE()
            stock_data = []

            for symbol in symbols:
                symbol_data = df_dict[df_dict["scripCode"] == symbol]

                if not symbol_data.empty:
                    stock_data.append(
                        {
                            "scripCode": symbol,
                            "companyName": symbol_data["companyName"].values[0],
                            sheet_name: symbol_data["weightedAvgPrice"].values[0],
                        }
                    )
                else:
                    print(f"No data found for symbol: {symbol}")
            df_stock = pd.DataFrame(stock_data)
            # st.write(df_stock)
            # st.write("Saving fetched data to Excel, new sheet name: sheet_name...")
            with pd.ExcelWriter(
                weightedAvgPrice_file_path, mode="a", engine="openpyxl"
            ) as writer:
                df_stock.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            st.error(f"Error: {e}")
        from openpyxl import load_workbook

        try:
            workbook = load_workbook("weightedAvgPrice.xlsx")
            source_sheet = workbook[sheet_name]
            workbook_all_sheets = load_workbook("All_sheets.xlsx")
            destination_sheet = workbook_all_sheets["weightedAvgPrice"]
            last_column_index = source_sheet.max_column
            last_column_values = []
            for row in source_sheet.iter_rows(
                min_row=1,
                max_row=source_sheet.max_row,
                min_col=last_column_index,
                max_col=last_column_index,
                values_only=True,
            ):
                for value in row:
                    last_column_values.append(value)
            next_column_index = destination_sheet.max_column + 1
            for i, value in enumerate(last_column_values, start=1):
                destination_sheet.cell(row=i, column=next_column_index, value=value)
            workbook_all_sheets.save("All_sheets.xlsx")
        except Exception as e:
            print(f"Error: {e}")

        # # totalTradedValue
        excel_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\stock_quote.xlsx"
        totalTradedValue_file_path = (
            r"C:\Users\HP\Staging\streamlit\stock_app\totalTradedValue.xlsx"
        )
        df_dict = pd.read_excel(excel_file_path, sheet_name=sheet_name)
        df_current = pd.read_excel(totalTradedValue_file_path)
        try:
            symbols = df_dict["scripCode"].tolist()
            b = BSE()
            stock_data = []

            for symbol in symbols:
                symbol_data = df_dict[df_dict["scripCode"] == symbol]

                if not symbol_data.empty:
                    stock_data.append(
                        {
                            "scripCode": symbol,
                            "companyName": symbol_data["companyName"].values[0],
                            sheet_name: symbol_data["totalTradedValue"].values[0],
                        }
                    )
                else:
                    print(f"No data found for symbol: {symbol}")
            df_stock = pd.DataFrame(stock_data)
            # st.write(df_stock)
            # st.write("Saving fetched data to Excel, new sheet name: sheet_name...")
            with pd.ExcelWriter(
                totalTradedValue_file_path, mode="a", engine="openpyxl"
            ) as writer:
                df_stock.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            st.error(f"Error: {e}")
        from openpyxl import load_workbook

        try:
            workbook = load_workbook("totalTradedValue.xlsx")
            source_sheet = workbook[sheet_name]
            workbook_all_sheets = load_workbook("All_sheets.xlsx")
            destination_sheet = workbook_all_sheets["totalTradedValue"]
            last_column_index = source_sheet.max_column
            last_column_values = []
            for row in source_sheet.iter_rows(
                min_row=1,
                max_row=source_sheet.max_row,
                min_col=last_column_index,
                max_col=last_column_index,
                values_only=True,
            ):
                for value in row:
                    last_column_values.append(value)
            next_column_index = destination_sheet.max_column + 1
            for i, value in enumerate(last_column_values, start=1):
                destination_sheet.cell(row=i, column=next_column_index, value=value)
            workbook_all_sheets.save("All_sheets.xlsx")
        except Exception as e:
            print(f"Error: {e}")

        # # totalTradedQuantity
        excel_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\stock_quote.xlsx"
        totalTradedQuantity_file_path = (
            r"C:\Users\HP\Staging\streamlit\stock_app\totalTradedQuantity.xlsx"
        )
        df_dict = pd.read_excel(excel_file_path, sheet_name=sheet_name)
        df_current = pd.read_excel(totalTradedQuantity_file_path)
        try:
            symbols = df_dict["scripCode"].tolist()
            b = BSE()
            stock_data = []

            for symbol in symbols:
                symbol_data = df_dict[df_dict["scripCode"] == symbol]

                if not symbol_data.empty:
                    stock_data.append(
                        {
                            "scripCode": symbol,
                            "companyName": symbol_data["companyName"].values[0],
                            sheet_name: symbol_data["totalTradedQuantity"].values[0],
                        }
                    )
                else:
                    print(f"No data found for symbol: {symbol}")
            df_stock = pd.DataFrame(stock_data)
            # st.write(df_stock)
            # st.write("Saving fetched data to Excel, new sheet name: sheet_name...")
            with pd.ExcelWriter(
                totalTradedQuantity_file_path, mode="a", engine="openpyxl"
            ) as writer:
                df_stock.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            st.error(f"Error: {e}")
        from openpyxl import load_workbook

        try:
            workbook = load_workbook("totalTradedQuantity.xlsx")
            source_sheet = workbook[sheet_name]
            workbook_all_sheets = load_workbook("All_sheets.xlsx")
            destination_sheet = workbook_all_sheets["totalTradedQuantity"]
            last_column_index = source_sheet.max_column
            last_column_values = []
            for row in source_sheet.iter_rows(
                min_row=1,
                max_row=source_sheet.max_row,
                min_col=last_column_index,
                max_col=last_column_index,
                values_only=True,
            ):
                for value in row:
                    last_column_values.append(value)
            next_column_index = destination_sheet.max_column + 1
            for i, value in enumerate(last_column_values, start=1):
                destination_sheet.cell(row=i, column=next_column_index, value=value)
            workbook_all_sheets.save("All_sheets.xlsx")
        except Exception as e:
            print(f"Error: {e}")

        # # 2WeekAvgQuantity
        excel_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\stock_quote.xlsx"
        twoWeekAvgQuantity_file_path = (
            r"C:\Users\HP\Staging\streamlit\stock_app\twoWeekAvgQuantity.xlsx"
        )
        df_dict = pd.read_excel(excel_file_path, sheet_name=sheet_name)
        df_current = pd.read_excel(twoWeekAvgQuantity_file_path)
        try:
            symbols = df_dict["scripCode"].tolist()
            b = BSE()
            stock_data = []

            for symbol in symbols:
                symbol_data = df_dict[df_dict["scripCode"] == symbol]

                if not symbol_data.empty:
                    stock_data.append(
                        {
                            "scripCode": symbol,
                            "companyName": symbol_data["companyName"].values[0],
                            sheet_name: symbol_data["2WeekAvgQuantity"].values[0],
                        }
                    )
                else:
                    print(f"No data found for symbol: {symbol}")
            df_stock = pd.DataFrame(stock_data)
            # st.write(df_stock)
            # st.write("Saving fetched data to Excel, new sheet name: sheet_name...")
            with pd.ExcelWriter(
                twoWeekAvgQuantity_file_path, mode="a", engine="openpyxl"
            ) as writer:
                df_stock.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            st.error(f"Error: {e}")
        from openpyxl import load_workbook

        try:
            workbook = load_workbook("twoWeekAvgQuantity.xlsx")
            source_sheet = workbook[sheet_name]
            workbook_all_sheets = load_workbook("All_sheets.xlsx")
            destination_sheet = workbook_all_sheets["2WeekAvgQuantity"]
            last_column_index = source_sheet.max_column
            last_column_values = []
            for row in source_sheet.iter_rows(
                min_row=1,
                max_row=source_sheet.max_row,
                min_col=last_column_index,
                max_col=last_column_index,
                values_only=True,
            ):
                for value in row:
                    last_column_values.append(value)
            next_column_index = destination_sheet.max_column + 1
            for i, value in enumerate(last_column_values, start=1):
                destination_sheet.cell(row=i, column=next_column_index, value=value)
            workbook_all_sheets.save("All_sheets.xlsx")
        except Exception as e:
            print(f"Error: {e}")

        # # marketCapFull
        excel_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\stock_quote.xlsx"
        marketCapFull_file_path = (
            r"C:\Users\HP\Staging\streamlit\stock_app\marketCapFull.xlsx"
        )
        df_dict = pd.read_excel(excel_file_path, sheet_name=sheet_name)
        df_current = pd.read_excel(marketCapFull_file_path)
        try:
            symbols = df_dict["scripCode"].tolist()
            b = BSE()
            stock_data = []

            for symbol in symbols:
                symbol_data = df_dict[df_dict["scripCode"] == symbol]

                if not symbol_data.empty:
                    stock_data.append(
                        {
                            "scripCode": symbol,
                            "companyName": symbol_data["companyName"].values[0],
                            sheet_name: symbol_data["marketCapFull"].values[0],
                        }
                    )
                else:
                    print(f"No data found for symbol: {symbol}")
            df_stock = pd.DataFrame(stock_data)
            # st.write(df_stock)
            # st.write("Saving fetched data to Excel, new sheet name: sheet_name...")
            with pd.ExcelWriter(
                marketCapFull_file_path, mode="a", engine="openpyxl"
            ) as writer:
                df_stock.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            st.error(f"Error: {e}")
        from openpyxl import load_workbook

        try:
            workbook = load_workbook("marketCapFull.xlsx")
            source_sheet = workbook[sheet_name]
            workbook_all_sheets = load_workbook("All_sheets.xlsx")
            destination_sheet = workbook_all_sheets["marketCapFull"]
            last_column_index = source_sheet.max_column
            last_column_values = []
            for row in source_sheet.iter_rows(
                min_row=1,
                max_row=source_sheet.max_row,
                min_col=last_column_index,
                max_col=last_column_index,
                values_only=True,
            ):
                for value in row:
                    last_column_values.append(value)
            next_column_index = destination_sheet.max_column + 1
            for i, value in enumerate(last_column_values, start=1):
                destination_sheet.cell(row=i, column=next_column_index, value=value)
            workbook_all_sheets.save("All_sheets.xlsx")
        except Exception as e:
            print(f"Error: {e}")

        # # marketCapFreeFloat
        excel_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\stock_quote.xlsx"
        marketCapFreeFloat_file_path = (
            r"C:\Users\HP\Staging\streamlit\stock_app\marketCapFreeFloat.xlsx"
        )
        df_dict = pd.read_excel(excel_file_path, sheet_name=sheet_name)
        df_current = pd.read_excel(marketCapFreeFloat_file_path)
        try:
            symbols = df_dict["scripCode"].tolist()
            b = BSE()
            stock_data = []

            for symbol in symbols:
                symbol_data = df_dict[df_dict["scripCode"] == symbol]

                if not symbol_data.empty:
                    stock_data.append(
                        {
                            "scripCode": symbol,
                            "companyName": symbol_data["companyName"].values[0],
                            sheet_name: symbol_data["marketCapFreeFloat"].values[0],
                        }
                    )
                else:
                    print(f"No data found for symbol: {symbol}")
            df_stock = pd.DataFrame(stock_data)
            # st.write(df_stock)
            # st.write("Saving fetched data to Excel, new sheet name: sheet_name...")
            with pd.ExcelWriter(
                marketCapFreeFloat_file_path, mode="a", engine="openpyxl"
            ) as writer:
                df_stock.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            st.error(f"Error: {e}")
        from openpyxl import load_workbook

        try:
            workbook = load_workbook("marketCapFreeFloat.xlsx")
            source_sheet = workbook[sheet_name]
            workbook_all_sheets = load_workbook("All_sheets.xlsx")
            destination_sheet = workbook_all_sheets["marketCapFreeFloat"]
            last_column_index = source_sheet.max_column
            last_column_values = []
            for row in source_sheet.iter_rows(
                min_row=1,
                max_row=source_sheet.max_row,
                min_col=last_column_index,
                max_col=last_column_index,
                values_only=True,
            ):
                for value in row:
                    last_column_values.append(value)
            next_column_index = destination_sheet.max_column + 1
            for i, value in enumerate(last_column_values, start=1):
                destination_sheet.cell(row=i, column=next_column_index, value=value)
            workbook_all_sheets.save("All_sheets.xlsx")
        except Exception as e:
            print(f"Error: {e}")

    if os.path.exists(r"C:\Users\HP\Staging\streamlit\stock_app\stock_quote.xlsx"):
        if st.button("Go to stock_quote.xlsx"):
            os.startfile(r"C:\Users\HP\Staging\streamlit\stock_app\stock_quote.xlsx")


if __name__ == "__main__":
    main()
