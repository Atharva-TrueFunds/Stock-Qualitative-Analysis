import numpy as np
import pandas as pd
from bsedata.bse import BSE
from datetime import datetime
import streamlit as st
from openpyxl import load_workbook

run_button_clicked = st.button("Run")
today = datetime.today().strftime("%d%b")
sheet_name = today + "_1"

excel_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\stock_quote.xlsx"
current_value_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\currentValue.xlsx"

df_dict = pd.read_excel(excel_file_path, sheet_name=sheet_name)
try:
    df_current = pd.read_excel(current_value_file_path)
except Exception as e:
    print(f"Error loading Excel file: {e}")
    with open("error.log", "a") as f:
        f.write(f"Error loading Excel file: {e}\n")
    st.error("Error loading Excel file. Please check the file path and sheet name.")


if run_button_clicked:
    try:
        symbols = df_dict["scripCode"].tolist()
        b = BSE()
        stock_data = []

        for symbol in symbols:
            symbol_data = df_dict[df_dict["scripCode"] == symbol]

            if not symbol_data.empty:
                stock_data.append(
                    {
                        "scripCode": symbol,
                        "companyName": symbol_data["companyName"].values[0],
                        sheet_name: symbol_data["currentValue"].values[0],
                    }
                )
            else:
                print(f"No data found for symbol: {symbol}")

        df_stock = pd.DataFrame(stock_data)

        st.write(df_stock)
        st.write("Saving fetched data to Excel, new sheet name: sheet_name...")

        with pd.ExcelWriter(
            current_value_file_path, mode="a", engine="openpyxl"
        ) as writer:
            df_stock.to_excel(writer, sheet_name=sheet_name, index=False)

    except Exception as e:
        st.error(f"Error: {e}")

    from openpyxl import load_workbook

    try:
        # Load the Excel workbook
        workbook = load_workbook("currentValue.xlsx")

        # Get references to the source and destination sheets
        source_sheet = workbook[sheet_name]
        destination_sheet = workbook["currentValue"]

        # Find the index of the last column in the source sheet
        last_column_index = source_sheet.max_column

        # Get the values from the last column of the source sheet
        last_column_values = []
        for row in source_sheet.iter_rows(
            min_row=1,
            max_row=source_sheet.max_row,
            min_col=last_column_index,
            max_col=last_column_index,
            values_only=True,
        ):
            for value in row:
                last_column_values.append(value)

        # Find the next available column in the destination sheet
        next_column_index = destination_sheet.max_column + 1

        # Write the values to the next available column in the destination sheet
        for i, value in enumerate(last_column_values, start=1):
            destination_sheet.cell(row=i, column=next_column_index, value=value)

        # Save the changes to the workbook
        workbook.save("currentValue.xlsx")

    except Exception as e:
        print(f"Error: {e}")
