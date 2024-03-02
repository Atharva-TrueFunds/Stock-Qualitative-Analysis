import os
import streamlit as st
from openpyxl import load_workbook

excel_file_path = r"C:\Users\HP\Staging\streamlit\stock_app\weightedAvgPrice.xlsx"
sheet_name = "weightedAvgPrice"

if os.path.exists(excel_file_path):
    if st.button("Go to weightedAvgPrice.xlsx"):
        try:
            workbook = load_workbook(excel_file_path)
            if sheet_name in workbook.sheetnames:
                os.startfile(excel_file_path)
            else:
                st.error(f"Sheet '{sheet_name}' not found in the Excel file.")
        except Exception as e:
            st.error(f"Error: {e}")
